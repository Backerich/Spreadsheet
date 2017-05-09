#!/usr/bin/env python3

import openpyxl

import sys
import os
import re
# import time
# import traceback


###############################################################################
# PROBLEME: -view unerwarteter error -> siehe unten (nicht wieder aufgetreten)
#           -
# IMPLIMENTIEREN: -make the cotinue func universal
#                 -übersicht
#                 -
###############################################################################


def exit():
    # Verlässt das Programm
    return sys.exit()


def clear():
    # Räumt den Terminal Output auf
    os.system("cls" if os.name == "nt" else "clear")


def continue_loop(wb, output):
    # Abfrage ob Programm beendet werden soll oder weiter laufen soll
    continue_loop = input(output)
    if (continue_loop == 'n') or (continue_loop == 'exit'): # quit Abfrage
        exit()
    else:
        # wenn nicht 'exit' geht es wieder zum Ausgangsmenü zurück
        menu(wb)

def what_sheet(wb):
    this_sheet = input("Welches Sheet willst du benutzen? \n>  ") 

    if this_sheet == 'exit':
        # quit Abfrage
        exit()
    elif this_sheet == "":
        # wenn nur Enter gedrückt wird wird das aktive Sheet als default genommen
        return wb.get_active_sheet()
    else: 
        try:
            # falls der Inputname existiert wird das ausgewählte Sheet benutzt
            return wb.get_sheet_by_name(this_sheet)
        except KeyError:
                # wenn nicht wird das Terminal aufgeräumt, eine Fehlermeldung ausgegeben und es wird erneut view abgefragt
                clear()
                print("\nDas Sheet {} existiert nicht. Versuchen sie es nochmal.\n".format(this_sheet))
                view(wb) 


def list_sheets(wb):
    # Listet alle Sheets in dem workbook auf
    try:
        # Absicherung durch unerwarteten Bug:
        # ERROR: AttributeError: 'NoneType' object has no attribute 'get_sheet_names'
        sheets = wb.get_sheet_names()
    except AttributeError:
        clear()
        print("Versuch es noch einmal.")
        menu(wb)
    counter = 1

    for string in sheets:
        print("Sheet Nummer {}: {}".format(counter, string))
        # counter zur Orientierung des Endusers
        counter += 1


def get_sheet(wb):
    # Zeigt alle Sheets, fragt ab welches benutzt werden soll und gibt dieses wieder
    list_sheets(wb)
    return what_sheet(wb)


def cell_value(sheet, row_first, column_first):
    # Ermittelt durch Koordinaten den Inhalt der ausgewählten Zelle
    return sheet.cell(row=int(row_first), column=int(column_first)).value


def all(wb):
    # Fragt neu gewünschtes Sheet ab
    sheet = get_sheet(wb)
    clear()

    # Ermittelt Maximale Reihe und Spalte
    row_count = sheet.max_row
    column_count = sheet.max_column
    # all_data = {}

    # TODO: Tabelle mit Werten des Sheets
    # header immer selbe row und anderer column
    for column in range(1 , column_count + 1):
        for row in range(1, row_count + 1):
            values = cell_value(sheet, row, column)
            if values != None:
                print(values)
                # all_data = {row , {}}
                print(all_data)
            else:
                print("\n")


def position(sheet):
    position_data = input("Geben sie die Positon ihrer Zelle an wie: 'Reihe, Spalte' (z.B. 2,1).\n>  ")

    if position_data != 'exit':
        clear()

        # RE: Muster zum filtern der Koordinaten 
        line = re.compile(r'''
        ^(?P<row>[\d]+),\s*
        (?P<column>[\d]+)$
        ''', re.X|re.M)

        # RE: Vergleicht das Muster mit dem Input und filtert je Reihe und Spalte heraus 
        # und gibt diese Werte and die 'cell_value' Methode weiter
        for match in line.finditer(position_data):
            return cell_value(sheet, match.group('row'), match.group('column'))
    else:
        exit()


def view(wb):
    # Fragt neu gewünschtes Sheet ab
    sheet = get_sheet(wb)
    clear()

    # Enthält den Inhalt der gewünschten Zelle
    position_data = position(sheet)

    if position_data != None:
        # Falls es einen Inhalt gibt:
        return "Ihre Zelle beinhaltet '{}'.".format(position_data)
    else:
        # Falls es keinen gibt:
        return "Diese Zelle gibt es nicht bitte versuchen sie es nochmal."


def ask_workbook():
    workbook_input = input("Gib deinen Pfad zur Datei an oder zieh ihn rein per drag and drop(es sollte eine Exel datei sein also z.B. .xlsx).\n>  ")
    if workbook_input == 'exit':
        exit()
    else:
        try:
            # Falls es den Path gibt verwendet er nun dieses Workbook
            return openpyxl.load_workbook(workbook_input)
        except FileNotFoundError:
            # Falls nicht wird eine Fehlermeldung ausgegeben und nochmal abgefragt
            clear()
            print("Diser Pfad exsistiert nicht. Versuch es nochmal.")
            ask_workbook()

def menu(wb):
    clear()
    while True:
        # Alle möglichen Functionen des Programmes als Menü
        print("Gib 'exit' ein um das Programm zu verlassen.")
        print("Gib 'view' ein um bestimmt Zeilen zu inspizieren.")
        print("Gib 'edit' ein um dein workbook zu ändern.")
        print("Gib 'all' ein um einen überblick von deinem Spread sheet zu erhalten.")
        user_input = input(">  ").lower()
        
        if user_input == 'exit':
            exit()
        elif user_input == 'view':
            # Falls Input 'view' Abfrage vom view um einzelnen Zelle Inhalte zu betrachen
            clear()
            print(view(wb))
            # Abrage ob das Programm beendet werden soll
            continue_loop(wb, "\nWollen sie weiter machen [Y/n]? \n>  ")
        elif user_input == 'edit':
            # Falls input 'edit' wird abgefragt ob der User das Workbook ändern will bzw. in welches
            clear()
            print("Ändere dein workbook.\n")
            # Ersetzt das aktuelle Workbook mit dem neuen
            wb = ask_workbook()
            clear()
        elif user_input == 'all':
            clear()
            # Zeigt alle Inhalte des Sheets an
            # TODO: Muss noch implimentiert werden
            all(wb)
            # Abrage ob das Programm beendet werden soll
            continue_loop(wb , "\nWollen sie weiter machen [Y/n]? \n>  ")
        else:
            # Fehlermeldung falls es diesen Menüpunkt nicht gibt und läuft die Schleife nochmal durch
            clear()
            print("Den Befehl '{}'' gibt es nicht. Versuchen sie es noch einmal.\n".format(user_input))
            continue
def main():
    clear()
    # DEBUG: Default Workbbok zum Debugen
    wb = openpyxl.load_workbook('Example/example_two.xlsx')
    # Fragt Workbook ab
    # wb = ask_workbook()
    menu(wb)

if __name__ == "__main__":
    main()




