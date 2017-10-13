#!/usr/bin/env python3

import openpyxl

import sys
import os
import locale

strings_language = []


class Clear(object):
    def exit(self):
        # Verlässt das Programm
        return sys.exit()

    def clear(self):
        # Räumt den Terminal Output auf
        os.system("cls" if os.name == "nt" else "clear")


class Something(object):
    def __init__(self):
        self.clear = Clear()
        # self.menu = Menu()
        
    def ask_workbook(self, input_string):
        while True:
            self.clear.clear()
            string = strings_language[6] + input_string + "\n" + "> "
    
            try:
                workbook_input = input(string).lower().strip()
            except KeyboardInterrupt:
                self.clear.exit()
    
            if workbook_input == strings_language[0]:
                self.clear.exit()
    
            else:
                try:
                    # Falls es den Path gibt verwendet er nun dieses Workbook
                    return openpyxl.load_workbook(workbook_input)
                except FileNotFoundError:
                    self.clear.clear()
                    ask_continue = input("This File does not exist. Do you want to continue? [Y/n]\n> ").lower().strip()
    
                    if ask_continue == 'n' or 'exit':
                        self.clear.exit()
                    else:
                        continue

    def continue_request(self, wb):
        # Abfrage ob Programm beendet werden soll oder weiter laufen soll
        try:
            self.continue_loop = input("\n" + strings_language[7] + "\n" + "> ").lower().strip()
        except KeyboardInterrupt:
            self.clear.exit()
    
        if (self.continue_loop == "n") or (self.continue_loop == strings_language[0]):
            self.clear.exit()
    
        else:
            # Wenn nicht 'exit' geht es wieder zum Ausgangsmenü zurück
            self.menu.menu(wb) # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! <- use this
            # self.clear.exit()
    
    def what_sheet(self, wb):
        try:
            self.this_sheet = input(strings_language[21] + "\n" + "> ").strip()
        except KeyboardInterrupt:
            self.clear.exit()
    
        if self.this_sheet == strings_language[0]:
            self.clear.exit()
    
        elif self.this_sheet == "":
            # Wenn nur Enter gedrückt wird wird das aktive Sheet als default genommen
            return wb.get_active_sheet()
    
        else:
            try:
                # Falls der Inputname existiert wird das ausgewählte Sheet benutzt
                return wb.get_sheet_by_name(self.this_sheet)
            except KeyError:
                # Wenn nicht wird das Terminal aufgeräumt, eine Fehlermeldung ausgegeben und es wird erneut view abgefragt
                self.clear.clear()
                print("\n" + strings_language[38] + "\n".format(self.this_sheet))
                self.get_sheet(wb)
    
    def list_sheets(self, wb):
        # Listet alle Sheets in dem workbook auf
        try:
            # Absicherung durch unerwarteten Bug:
            # ERROR: AttributeError: 'NoneType' object has no attribute 'get_sheet_names'
            self.sheets = wb.get_sheet_names()
        except AttributeError:
            self.clear.clear()
            print(strings_language[20])
            # self.menu.menu(wb) !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!! <- use this
            self.clear.exit()
        counter = 1
    
        for string in self.sheets:
            print(strings_language[37].format(counter, string))
    
            # counter zur Orientierung des Endusers
            counter += 1
    
    def max_sheet(self, sheet):
        try:
            # Falls es eine Reihe und Spalte gibt wird die höchste wiedergegeben
            return sheet.max_row, sheet.max_column
        except AttributeError:
            # Falls es keine Reihe oder Spalte gibt wird eine Fahlermeldung ausgegeben
            print(strings_language[19])
    
    def get_sheet(self, wb):
        # Zeigt alle Sheets, fragt ab welches benutzt werden soll und gibt dieses wieder
        self.list_sheets(wb)
        return self.what_sheet(wb)
    
    def cell_value(self, sheet, row_first, column_first):
        # Ermittelt durch Koordinaten den Inhalt der ausgewählten Zelle
        return sheet.cell(row=int(row_first), column=int(column_first))
    
    def get_values(self, sheet, row_count, column_count):
        all_raw_rows = []
        longest = 0
    
        # Fügt die Spalten beschreibung hinzu
        all_raw_rows.append(i for i in range(1, row_count + 1))
    
        for row in range(1, row_count + 1):
            innerlist = []
    
            for column in range(1, column_count + 1):
                # Ermittelt Zellen Inhalt
                values = self.cell_value(sheet, row, column).value
    
                if values != None:
                    innerlist.append(values)
    
                    # Ermittelt längsten Wert
                    if len(values) > longest:
                        longest = len(values)
    
                else:
                    # Abgrenzungen
                    innerlist.append("")
    
            # Fügt alle zusammen
            all_raw_rows.append(innerlist)
        return all_raw_rows, longest
    
    def copy(self):
        # Fragt nach dem Namen der Datei C
        print(strings_language[18])
        try:
            self.copy_name = input(strings_language[17] + "\n" + "> ").lower().strip()
        except KeyboardInterrupt:
            self.clear.exit()
    
        if self.copy_name == strings_language[0]:
            self.clear.exit()
    
        else:
    
            # Gibt den Pfad zurück
            return "Example/" + self.copy_name

class Compare(object):
    def __init__(self):
        self.clear = Clear()
        self.som = Something()

    def compare_sheets(self, first_values, second_values, first_data, first_workbook):
        first_sheet_values = []
    
        for first_row in first_values:
            for first_item in first_row:
                if first_item != "":
    
                    # Teilt die Ausgangszelle per Leerzeichen
                    item_list = first_item.split(" ")  # RE eher angebracht
    
                    for item_raw in item_list:
                        # Filtert Kommata heraus
                        item = item_raw.replace(",", "")  # RE eher angebracht
    
                        # Ermittelt die Postion der Reihe, der Spalte und die Position im String
                        self.position_row = first_values.index(first_row)
                        self.position_column = first_row.index(first_item)
                        self.position_string = item_list.index(item_raw)
    
                        # Gibt das Objekt mit den Koordinaten weiter
                        first_sheet_values.append([item, (self.position_row + 1, self.position_column + 1, self.position_string)])
    
                    # Setzt "done" ans ende des benutzten Objektes damit selbes Item mit anderen Koordinaten existieren kann
                    first_row[self.position_column] += "done"
    
        for second_row in second_values:
            for second_item in second_row:
    
                # Wenn item Leer ist weiter machen
                if second_item == "":
                    pass
    
                else:
                    # Löscht Leerzeichen und teilt Werte auf durch "="
                    second_item = second_item.replace(" ", "").split("=")
    
                    for cell in first_sheet_values:  # iter nicht immer über alle
                        item = cell[0]
    
                        # Wenn Übereinstimmung des Wertes und des Vergleichsobjekt
                        if second_item[0] == item:
    
                            # Ermittelt die position
                            row = cell[1][0]
                            column = cell[1][1]
                            pos = self.som.cell_value(first_data, row, column).coordinate
    
                            # Ermittelt Inhalt der Zelle
                            value = first_data[pos].value
    
                            ####################################################################
                            # Gucken ob 'find' nicht den oberen Abschnitt der Funktion ersetzt #
                            ####################################################################
    
                            # Sucht Vergleichs datei in Ausgangsdatei
                            index = value.find(second_item[0])
    
                            try:  # Checken ob Errorhandling nötig
                                # Löscht Vergleichsdatei in Ausgangsdatie
                                value = value.replace(second_item[0], "")
    
                                # Setzt String wieder zusammen
                                value = value[:index] + second_item[1] + value[index:]
    
                                first_data[pos] = value
                            except IndexError:
                                pass
        # Erstellt kopie von von Datei A und speichert es in Datei C
        third_name = self.som.copy()
        first_workbook.save(third_name)
    
    
    def sheets_to_compare(self, wb):  # Überlegung ob nötig
        try:
            self.user_workbook = input(strings_language[16] + "\n" + "> ").lower().strip()
        except KeyboardInterrupt:
            self.clear.exit()
    
        sheet = None
        workbook = None
    
        if self.user_workbook == strings_language[0]:
            self.clear.exit()
    
        elif self.user_workbook == "n":
            # Erfolgt wenn anderes Workbook genutzt werden soll und ermittelt das Worksheet
            workbook = self.som.ask_workbook("\n" + strings_language[15])
            self.clear.clear()
            self.som.get_sheet(workbook)
            self.clear.clear()
    
        else:
            # Erfolgt wenn selbes Workbook genutzt werden soll und ermittelt das Worksheet
            workbook = wb
            self.clear.clear()
            self.sheet = self.som.get_sheet(workbook)
            self.clear.clear()
    
        return self.sheet, self.user_workbook, workbook

class Grid(object):
    def __init__(self):
        self.clear = Clear()
        self.som = Something()

    def position(self, sheet):
        while True:
            try:
                self.position_data = input(strings_language[12] + "\n" + "> ").lower().strip()
            except KeyboardInterrupt:
                self.clear.exit()
    
            if self.position_data == strings_language[0]:
                self.clear.exit()
    
            else:
                self.clear.clear()
                self.position_data = self.position_data.split(",")
    
                try:
                    strings_language[36].format(self.position_data)
                    print(strings_language[36].format(self.som.cell_value(sheet, self.position_data[0], self.position_data[1]).value))
    
                except (ValueError, IndexError):
                    self.clear.clear()
                    ask_continue = input("This Position does not exist. Do you want to continue? [Y/n]\n> ").lower().strip()
    
                    if ask_continue == 'n' or 'exit':
                        self.clear.exit()
                    else:
                        continue
    
    
    def grid(self, values, longest):
        # Convertiert die Werte in symetrische Reihen aus Strings
        string_rows = []
        for rows_in_list in values:
            raw_temp = []
    
            for item in rows_in_list:
                # Ermittelt die länge der spezifischen Zelle
                value_length = longest + 1 - len(str(item))
    
                # Setzt die Reihe zusammen
                raw_temp.append(str(item) + " " * value_length + "|")
    
            # Fügt die Reihe dem Raster hinzu
            string_rows.append(raw_temp)
    
        # Fügt Reihenzahlen hinzu und gibt die Tabelle aus
        for row in string_rows:
            string_temp = ""
    
            # Ermittelt die Reihenzahl
            string_temp += str(string_rows.index(row)) + "| "
    
            for i in range(0, len(row)):
                # Fügt die Reihenzahlen und die restliche Reihe zusammen
                string_temp += row[i]
    
            # Gibt die Tabelle aus
            print(string_temp)


class Menu(object):
    def __init__(self):
        self.clear = Clear()
        self.compare = Compare()
        self.grid = Grid()

        self.som = Something()

    def compares(self, wb):
        # Ermittelt die gewünschten Worksheets und Workbooks
        # Erstes Worksheet
        data_first = self.compare.sheets_to_compare(wb)
        first_data = data_first[0]
        first_workbook_name = data_first[1]
        first_workbook = data_first[2]

        # Zweites Worksheet
        data_second = self.compare.sheets_to_compare(wb)  # BUG: Falsche Strings
        second_data = data_second[0]

        # Ermittelt Maximale Reihe und Spalte des ersten Worksheets
        first_max_sheet = self.som.max_sheet(first_data)
        first_max_row = first_max_sheet[0]
        first_max_column = first_max_sheet[1]

        # Ermittelt Maximale Reihe und Spalte des zweiten Worksheets
        second_max_sheet = self.som.max_sheet(second_data)
        second_max_row = second_max_sheet[0]
        second_max_column = second_max_sheet[1]

        # Ermittelt die Werte der Worksheets
        first_values = self.som.get_values(first_data, first_max_row, first_max_column)[0]
        second_values = self.som.get_values(second_data, second_max_row, second_max_column)[0]

        # Löschen des ersten Wertes
        del first_values[0]
        del second_values[0]

        # Vergleicht die Worksheets und gibt die Ergebnisse aus
        self.compare.compare_sheets(first_values, second_values, first_data, first_workbook)

    def all(self, wb):
        # Fragt das Worksheet ab
        sheet = self.som.get_sheet(wb)
        self.clear.clear()

        # Ermittelt Maximale Reihe und Spalte
        max_row_column = self.som.max_sheet(sheet)
        max_row = max_row_column[0]
        max_column = max_row_column[1]

        # Ermittelt die Rohwerte der Reihen
        values = self.som.get_values(sheet, max_row, max_column)[0]

        # Ermittelt die längste Zelle des Worksheets
        longest = self.som.get_values(sheet, max_row, max_column)[1]

        # Verarbeitet die Werte zu einem Grid von Strings
        self.grid.grid(values, longest)

    def view(self, wb):
        # Fragt das Worksheet ab
        sheet = self.som.get_sheet(wb)
        self.clear.clear()

        # Ermittelt den Wert der Zelle und gibt ihn aus
        self.grid.position(sheet)

    def help(self):
        # Die einzelnen Funktionen und ihre Beschreibung
        funktionen = {
            strings_language[5]: strings_language[31],
            strings_language[4]: strings_language[32],
            strings_language[3]: strings_language[33],
            strings_language[0]: strings_language[34],
            strings_language[30]: "https://github.com/Backerich/Spreadsheet"}

        # Ermittelt den längsten Key
        longest_key = 0
        for key, value in funktionen.items():
            if len(key) > longest_key:
                longest_key = len(key)

        # Gibt die Tabelle aus mit den Funktionsnamen und Beschreibungen
        for key, value in funktionen.items():
            print(key + " " * (longest_key - len(key)) + "| " + value)

        # Fragt nach der Funktion die weiter Erläutert werden soll
        try:
            self.ask_function = input("\n" + strings_language[10] + "\n" + "> ").lower().strip()
        except KeyboardInterrupt:
            self.clear.exit()

        if self.ask_function == strings_language[0]:
            self.clear.exit()

        else:
            # Gibt theoretisch die weitere Beschreibung aus
            self.clear.clear()
            print(strings_language[9])

    def menu(self, wb):
        # Das Ausgangsmenü um auf alle Funktionen zugreifen zu können
        self.clear.clear()

        while True:
            # Alle möglichen Funktionen des Programmes
            print(strings_language[24].format(strings_language[5]))
            print(strings_language[25].format(strings_language[4]))
            print(strings_language[26].format(strings_language[3]))
            print(strings_language[27].format(strings_language[2]))
            print(strings_language[28].format(strings_language[1]))
            print(strings_language[29].format(strings_language[0]))

            # Benutzerinput für das Menü
            try:
                user_input = input("> ").lower().strip()
            except KeyboardInterrupt:
                self.clear.exit()

            # DEBUG: Als default der Eingabe:
            # user_input = strings_language[3]

            if user_input == strings_language[0]:
                self.clear.exit()

            elif user_input == strings_language[4]:
                # Zum betrachten des Wertes einer einzelner Zellen
                self.clear.clear()
                self.view(wb)
                self.som.continue_request(wb)

            elif user_input == strings_language[2]:
                # Zum ändern des aktuellen workbooks
                self.clear.clear()
                print(strings_language[8] + "\n")
                wb = self.som.ask_workbook("")
                self.clear.clear()

            elif user_input == strings_language[5]:
                # Gibt den gesammten Inhalt des Sheets in einer Tabelle aus
                self.clear.clear()
                self.all(wb)
                self.som.continue_request(wb)

            elif user_input == strings_language[3]:
                # Datei A wird mit Datei B verglichen und einstimmige Komponenten in Datei C geschrieben und ersetzt
                self.clear.clear()
                self.compares(wb)
                self.som.continue_request(wb)

            elif user_input == strings_language[1]:
                # Beschreibt die Funktionen des Programmes
                self.clear.clear()
                self.help()
                self.som.continue_request(wb)

            else:
                # ERROR: Wenn es die Funktion nicht gibt
                self.clear.clear()
                print(strings_language[23].format(user_input) + "\n")
                continue

    def language(self):
        locale.getdefaultlocale()

        while True:
            self.clear.clear()
            print("Type 'ger' for german.")
            print("Type 'eng' for englisch.")

            try:
                self.ask_language = input("Which language should your programm use?\n> ").lower().strip()
            except KeyboardInterrupt:
                self.clear.exit()

            if self.ask_language == 'exit':
                self.clear.exit()

            elif self.ask_language == 'ger':
                # Liest die Deutschen Strings aus
                with open("Languages/german", "r") as file:
                    for line in file:
                        full_line = line.replace("\n", "")
                        strings_language.append(full_line)
                break

            elif self.ask_language == 'eng':
                # Liest die Englischen Strings aus
                with open("Languages/englisch", "r") as file:
                    for line in file:
                        full_line = line.replace("\n", "")
                        strings_language.append(full_line)
                break

            else:
                self.clear.clear()
                ask_continue = input("This language does not exist. Do you want to continue? [Y/n]\n> ").lower().strip()
                if ask_continue != 'exit' or 'n':
                    continue
                else:
                    self.clear.exit()


def main():
    menu = Menu()
    clear = Clear()

    som = Something()

    # Fragt Welche Sprache
    menu.language()
    clear.clear()

    # DEBUG: Default Workbbok zum Debuggen
    wb = openpyxl.load_workbook('Example/example.xlsx')

    # Fragt nach dem workbook
    # wb = som.ask_workbook("")  # BUG: Wenn falsch kann compare nicht weiter arbeiten

    # Ruft das Menü auf mit allen Funktionen des Programmes
    menu.menu(wb)


if __name__ == "__main__":
    main()

